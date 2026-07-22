from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
import re
from typing import Iterable

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPORTS_REQUIRED_COLUMNS = [
    "ProductsWrongInstructions_ID",
    "Kód produktu",
    "StoreJob_ID",
    "Přeložené instrukce",
    "Jméno uživatele",
    "Pobočka",
    "Host",
    "Vytvořeno",
    "Zkontrolováno",
    "Zkontrolováno logistikou datum",
    "Název stanice",
]


PRODUCT_EXPECTED_COLUMNS = [
    "Kód produktu",
    "Název produktu",
    "Segmentace (obecná)",
    "Segment1",
    "Segment2",
    "Segment3",
    "SEOPrefix_ID",
    "SC-Skupina",
    "Typ obalu (Logistické parametry)",
    "Výška",
    "Šířka",
    "Hloubka",
    "Váha",
    "Size1",
    "Size2",
    "Size3",
    "GeoSize (id)",
    "Dobalovat (Skladová vlastnost)",
    "Nebalit (Skladová vlastnost)",
    "Bublinková fólie (Skladová vlastnost)",
    "Fólie (Skladová vlastnost)",
    "Obálka (Skladová vlastnost)",
    "Kartonová krabička (Skladová vlastnost)",
    "Prelep páskou (Skladová vlastnost)",
    "Přelep uzávěr (Skladová vlastnost)",
    "Utěsnit uzávěr (Skladová vlastnost)",
    "IsFragile",
    "Lehko poškoditelný (Logistické parametry)",
    "Křehký produkt (Logistické parametry)",
    "Často poškozovaný / reklamovaný (Skladová vlastnost)",
    "Sklo (Logistické parametry)",
    "Porcelán (Logistické parametry)",
    "Teflon (Logistické parametry)",
    "Ostré hrany (Logistické parametry)",
    "Tenká lepenka (lego) (Logistické parametry)",
    "Lehko znečistitelný (Logistické parametry)",
    "Znečistí okolí (Logistické parametry)",
    "Tekutý (Logistické parametry)",
    "Má šroubovací uzávěr (Logistické parametry)",
    "Má lehce oddělitelný uzávěr (Logistické parametry)",
    "Sběratelská edice (Logistické parametry)",
    "Shockwatch (Skladová vlastnost)",
]


RISK_COLUMNS = [
    "IsFragile",
    "Lehko poškoditelný (Logistické parametry)",
    "Křehký produkt (Logistické parametry)",
    "Často poškozovaný / reklamovaný (Skladová vlastnost)",
    "Sklo (Logistické parametry)",
    "Porcelán (Logistické parametry)",
    "Teflon (Logistické parametry)",
    "Ostré hrany (Logistické parametry)",
    "Tenká lepenka (lego) (Logistické parametry)",
    "Lehko znečistitelný (Logistické parametry)",
    "Znečistí okolí (Logistické parametry)",
    "Tekutý (Logistické parametry)",
    "Má šroubovací uzávěr (Logistické parametry)",
    "Má lehce oddělitelný uzávěr (Logistické parametry)",
    "Sběratelská edice (Logistické parametry)",
    "Shockwatch (Skladová vlastnost)",
]


INSTRUCTION_CATEGORY_ORDER = [
    ("Bublinková fólie", ["bublink"]),
    ("Kartonová krabička", ["karton", "krabi"]),
    ("Obálka", ["obál", "obalk"]),
    ("Prelep páskou", ["pásk", "pask"]),
    ("Utěsnit uzávěr", ["utěsni", "utesni"]),
    ("Přelep uzávěr", ["přelep", "prelep"]),
    ("Fólie", ["fóli", "foli"]),
    ("Štítok", ["štítek", "stitek"]),
]


TRIGGER_PARAMETER_BY_CATEGORY = {
    "Bublinková fólie": "Bublinková fólie (Skladová vlastnost)",
    "Fólie": "Fólie (Skladová vlastnost)",
    "Obálka": "Obálka (Skladová vlastnost)",
    "Kartonová krabička": "Kartonová krabička (Skladová vlastnost)",
    "Prelep páskou": "Prelep páskou (Skladová vlastnost)",
    "Přelep uzávěr": "Přelep uzávěr (Skladová vlastnost)",
    "Utěsnit uzávěr": "Utěsnit uzávěr (Skladová vlastnost)",
}


TRUE_VALUES = {
    "1",
    "true",
    "pravda",
    "ano",
    "áno",
    "yes",
    "y",
}

FALSE_VALUES = {
    "0",
    "false",
    "nepravda",
    "ne",
    "nie",
    "no",
}


LONG_TEXT_COLUMN_WIDTHS = {
    "decision_reason": 80,
    "validator_checklist": 80,
    "active_risk_parameters": 70,
    "ai_decision_hint": 80,
    "Přeložené instrukce": 70,
    "PĹ™eloĹľenĂ© instrukce": 70,
}


SECTION_FILL = PatternFill(fill_type="solid", fgColor="D9EAD3")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E2F3")


def _strip_and_collapse_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def normalize_text_value(value) -> str | None:
    if pd.isna(value):
        return None
    if isinstance(value, str):
        cleaned = _strip_and_collapse_spaces(value)
        return cleaned if cleaned != "" else None
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int,)) and not isinstance(value, bool):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    cleaned = _strip_and_collapse_spaces(str(value))
    return cleaned if cleaned != "" else None


def clean_code_value(value) -> str | None:
    text = normalize_text_value(value)
    if text is None:
        return None
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def ensure_column(df: pd.DataFrame, column: str, default=None) -> pd.DataFrame:
    if column not in df.columns:
        df[column] = default
    return df


def warn_missing_columns(df: pd.DataFrame, expected: Iterable[str]) -> list[str]:
    missing = [column for column in expected if column not in df.columns]
    return missing


def prepare_reports_clean(raw_reports: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    warnings: list[str] = []
    df = raw_reports.copy()

    missing = warn_missing_columns(df, REPORTS_REQUIRED_COLUMNS)
    if missing:
        warnings.append(
            "Missing expected report columns: " + ", ".join(missing)
        )

    for column in REPORTS_REQUIRED_COLUMNS:
        ensure_column(df, column, None)

    rename_map = {"ProductsWrongInstructions_ID": "report_id"}
    df = df.rename(columns=rename_map)

    for column in [
        "Kód produktu",
        "Přeložené instrukce",
        "Jméno uživatele",
        "Pobočka",
        "Název stanice",
        "StoreJob_ID",
    ]:
        df[column] = df[column].map(normalize_text_value)

    df["Kód produktu"] = df["Kód produktu"].map(clean_code_value).astype("string")
    df["Vytvořeno"] = pd.to_datetime(df["Vytvořeno"], errors="coerce")

    ordered_columns = [
        "report_id",
        "Kód produktu",
        "StoreJob_ID",
        "Přeložené instrukce",
        "Jméno uživatele",
        "Pobočka",
        "Host",
        "Vytvořeno",
        "Zkontrolováno",
        "Zkontrolováno logistikou datum",
        "Název stanice",
    ]
    reports_clean = df.loc[:, ordered_columns].copy()
    return reports_clean, warnings


def prepare_products_clean(raw_products: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    warnings: list[str] = []
    df = raw_products.copy()

    missing = warn_missing_columns(df, PRODUCT_EXPECTED_COLUMNS)
    if missing:
        warnings.append(
            "Missing expected product columns: " + ", ".join(missing)
        )

    ensure_column(df, "Kód produktu", None)
    df["Kód produktu"] = df["Kód produktu"].map(clean_code_value).astype("string")

    return df, warnings


def build_invalid_reports(reports_clean: pd.DataFrame) -> pd.DataFrame:
    df = reports_clean.copy()
    required_fields = ["report_id", "KĂłd produktu", "PĹ™eloĹľenĂ© instrukce"]
    available_fields = [column for column in required_fields if column in df.columns]
    if not available_fields:
        invalid = df.head(0).copy()
        invalid["invalid_reason"] = pd.Series(dtype="string")
        return invalid

    invalid_mask = df[available_fields].isna().any(axis=1)
    invalid = df.loc[invalid_mask].copy()

    def _reason(row: pd.Series) -> str:
        missing = [column for column in required_fields if column in row.index and pd.isna(row[column])]
        return "; ".join(missing) if missing else "unknown"

    invalid["invalid_reason"] = invalid.apply(_reason, axis=1)
    return invalid.reset_index(drop=True)


def group_cases_mvp(reports_clean: pd.DataFrame) -> pd.DataFrame:
    df = reports_clean.copy()
    ensure_column(df, "Kód produktu", None)
    ensure_column(df, "Přeložené instrukce", None)

    grouped = (
        df.groupby(["Kód produktu", "Přeložené instrukce"], dropna=False, sort=False)
        .agg(
            report_count=("report_id", "size"),
            unique_users_count=("Jméno uživatele", pd.Series.nunique),
            first_reported_at=("Vytvořeno", "min"),
            last_reported_at=("Vytvořeno", "max"),
            branch_count=("Pobočka", pd.Series.nunique),
            station_count=("Název stanice", pd.Series.nunique),
            sample_storejob_ids=("StoreJob_ID", lambda s: ", ".join(_distinct_first_values(s, 5))),
        )
        .reset_index()
    )

    grouped = grouped.sort_values(
        by=["report_count", "last_reported_at"],
        ascending=[False, False],
        kind="mergesort",
    ).reset_index(drop=True)

    grouped.insert(0, "case_number", range(1, len(grouped) + 1))
    grouped.insert(0, "case_id", [f"CI-{i:06d}" for i in range(1, len(grouped) + 1)])
    grouped["status"] = "Neriešené"
    grouped["priority"] = grouped["report_count"].map(priority_from_count)
    grouped["owner"] = ""
    grouped["validator_note"] = ""

    return grouped[
        [
            "case_id",
            "case_number",
            "Kód produktu",
            "Přeložené instrukce",
            "report_count",
            "unique_users_count",
            "first_reported_at",
            "last_reported_at",
            "branch_count",
            "station_count",
            "sample_storejob_ids",
            "status",
            "priority",
            "owner",
            "validator_note",
        ]
    ]


def _distinct_first_values(values: pd.Series, limit: int) -> list[str]:
    seen: list[str] = []
    for value in values:
        text = normalize_text_value(value)
        if text and text not in seen:
            seen.append(text)
        if len(seen) >= limit:
            break
    return seen


def priority_from_count(report_count: int) -> str:
    if report_count >= 20:
        return "Vysoká"
    if report_count >= 5:
        return "Stredná"
    return "Nízka"


def classify_instruction(text) -> str:
    normalized = normalize_text_value(text)
    if not normalized:
        return "Iné"
    lowered = normalized.lower()
    for category, tokens in INSTRUCTION_CATEGORY_ORDER:
        if any(token in lowered for token in tokens):
            return category
    return "Iné"


def main_trigger_parameter(category: str) -> str:
    return TRIGGER_PARAMETER_BY_CATEGORY.get(category, "Nezistené")


def is_active(value) -> bool:
    if pd.isna(value):
        return False
    if isinstance(value, bool):
        return value
    text = normalize_text_value(value)
    if text is None:
        return False
    lowered = text.lower()
    if lowered in TRUE_VALUES:
        return True
    if lowered in FALSE_VALUES:
        return False
    return False


def build_cases_enriched(cases_mvp: pd.DataFrame, products_clean: pd.DataFrame) -> pd.DataFrame:
    cases = cases_mvp.copy()
    products = products_clean.copy()

    if "Kód produktu" not in products.columns:
        products["Kód produktu"] = None

    cases["Kód produktu"] = cases["Kód produktu"].map(clean_code_value).astype("string")
    products["Kód produktu"] = products["Kód produktu"].map(clean_code_value).astype("string")

    merged = cases.merge(
        products,
        how="left",
        on="Kód produktu",
        suffixes=("", "_product"),
    )

    if "Název produktu" not in merged.columns:
        merged["Název produktu"] = pd.NA

    merged["product_match_status"] = merged["Název produktu"].apply(
        lambda value: "Produkt nájdený" if not pd.isna(value) and normalize_text_value(value) else "Produkt nenájdený"
    )

    merged["instruction_category"] = merged["Přeložené instrukce"].apply(classify_instruction)
    merged["main_trigger_parameter"] = merged["instruction_category"].apply(main_trigger_parameter)

    def _trigger_value(row):
        column = row["main_trigger_parameter"]
        if column == "Nezistené":
            return pd.NA
        if column not in row.index:
            return pd.NA
        return row[column]

    merged["trigger_parameter_value"] = merged.apply(_trigger_value, axis=1)
    merged["trigger_detected"] = merged["trigger_parameter_value"].apply(lambda value: 1 if is_active(value) else 0)

    def _active_risks(row):
        active = []
        for column in RISK_COLUMNS:
            if column in row.index and is_active(row[column]):
                active.append(column)
        if not active:
            return "Žiadne aktívne rizikové parametre"
        return "; ".join(active)

    merged["active_risk_parameters"] = merged.apply(_active_risks, axis=1)
    merged["has_risk_flag"] = merged["active_risk_parameters"].apply(
        lambda value: 0 if value == "Žiadne aktívne rizikové parametre" else 1
    )
    merged["risk_evaluation"] = merged.apply(
        lambda row: "Produkt nemá aktívne rizikové parametre v dostupných dátach."
        if row["has_risk_flag"] == 0
        else f"Produkt má aktívne rizikové parametre: {row['active_risk_parameters']}. Balenie môže byť oprávnené a musí byť validované opatrne.",
        axis=1,
    )

    def _decision_hint(row):
        if row["product_match_status"] != "Produkt nájdený":
            return "Produkt sa nepodarilo napojiť na produktové vlastnosti. Odporúčaná manuálna analýza."
        if row["trigger_detected"] == 1 and row["has_risk_flag"] == 0:
            return (
                f"Trigger parameter bol nájdený: {row['main_trigger_parameter']} = aktívny. "
                "Produkt nemá aktívne rizikové parametre. Preveriť, či je tento baliaci parameter nastavený oprávnene."
            )
        if row["trigger_detected"] == 1 and row["has_risk_flag"] == 1:
            return (
                f"Trigger parameter bol nájdený: {row['main_trigger_parameter']} = aktívny. "
                f"Produkt má rizikové parametre: {row['active_risk_parameters']}. Balenie môže byť oprávnené."
            )
        if row["trigger_detected"] == 0 and row["main_trigger_parameter"] == "Nezistené":
            return "Trigger parameter nie je známy. Preveriť systémovú logiku alebo pravidlo zobrazovania inštrukcie."
        return "Dáta sú nejednoznačné. Odporúčaná manuálna analýza."

    merged["ai_decision_hint"] = merged.apply(_decision_hint, axis=1)

    return merged


def build_decision_maker(cases_enriched: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    df = cases_enriched.copy()
    warnings: list[str] = []

    required_columns = [
        "product_match_status",
        "trigger_detected",
        "has_risk_flag",
        "main_trigger_parameter",
        "instruction_category",
    ]
    missing = warn_missing_columns(df, required_columns)
    if missing:
        warnings.append(
            "Missing expected decision columns: " + ", ".join(missing)
        )

    def _row_text(row: pd.Series, column: str) -> str | None:
        if column not in row.index:
            return None
        return normalize_text_value(row[column])

    def _row_int(row: pd.Series, column: str) -> int | None:
        if column not in row.index or pd.isna(row[column]):
            return None
        try:
            return int(row[column])
        except (TypeError, ValueError):
            text = normalize_text_value(row[column])
            if text is None:
                return None
            try:
                return int(float(text))
            except (TypeError, ValueError):
                return None

    def _rule_values(row: pd.Series) -> tuple[str, str, str, str, str]:
        product_match_status = _row_text(row, "product_match_status")
        trigger_detected = _row_int(row, "trigger_detected")
        has_risk_flag = _row_int(row, "has_risk_flag")
        main_trigger_parameter = _row_text(row, "main_trigger_parameter")
        if not main_trigger_parameter:
            main_trigger_parameter = "Nezistené"

        if product_match_status != "Produkt nájdený":
            return (
                "Manuálna analýza",
                "Kód produktu / produktový master",
                "Produkt sa nepodarilo napojiť na produktové vlastnosti. Bez produktových dát nie je možné bezpečne rozhodnúť.",
                "nízka",
                "Overiť Kód produktu; overiť dostupnosť produktu v produktovom masteri; doplniť produktové dáta alebo vyradiť neplatný report.",
            )

        if main_trigger_parameter == "Nezistené":
            return (
                "Manuálna analýza",
                "instruction_category",
                "Kategóriu inštrukcie sa nepodarilo jednoznačne namapovať na známy trigger parameter.",
                "nízka",
                "Prečítať text inštrukcie; zaradiť ju do správnej kategórie; doplniť mapovanie instruction_category; overiť, či nejde o nový typ baliacej inštrukcie.",
            )

        if trigger_detected is None or has_risk_flag is None:
            return (
                "Manuálna analýza",
                main_trigger_parameter,
                "Dáta sú nejednoznačné a nebolo možné použiť žiadne pravidlo Decision Makeru.",
                "nízka",
                "Manuálne preveriť produkt, inštrukciu, trigger parameter a rizikové vlastnosti.",
            )

        if trigger_detected == 1 and has_risk_flag == 0:
            return (
                "Preveriť odstránenie balenia",
                main_trigger_parameter,
                "Inštrukcia má aktívny systémový trigger, ale produkt nemá aktívne rizikové parametre v dostupných dátach. Treba preveriť, či je baliaci parameter nastavený oprávnene.",
                "stredná",
                "Fyzicky overiť obal produktu; skontrolovať poškodenosť/reklamovanosť; porovnať podobné produkty v segmente; ak nie je dôvod na balenie, navrhnúť vypnutie trigger parametra.",
            )

        if trigger_detected == 1 and has_risk_flag == 1:
            return (
                "Ponechať balenie",
                main_trigger_parameter,
                "Inštrukcia má aktívny systémový trigger a produkt má aktívne rizikové parametre. Balenie môže byť oprávnené.",
                "stredná",
                "Overiť, či aktívne rizikové parametre skutočne zodpovedajú fyzickému produktu; skontrolovať, či zvolený spôsob balenia je primeraný; pri pochybnosti porovnať podobné produkty.",
            )

        if trigger_detected == 0 and main_trigger_parameter != "Nezistené":
            return (
                "Odovzdať na IT",
                main_trigger_parameter,
                "Inštrukcia sa zobrazuje, ale očakávaný trigger parameter nie je aktívny. Môže ísť o inú systémovú logiku, pravidlo zobrazovania, cache alebo chybné mapovanie.",
                "stredná",
                "Overiť hodnotu trigger parametra; skontrolovať baliace pravidlá; overiť, či inštrukcia nevzniká z iného pravidla; pripraviť príklad pre IT vrátane case_id, SKU a StoreJob_ID.",
            )

        return (
            "Manuálna analýza",
            "Nezistené",
            "Dáta sú nejednoznačné a nebolo možné použiť žiadne pravidlo Decision Makeru.",
            "nízka",
            "Manuálne preveriť produkt, inštrukciu, trigger parameter a rizikové vlastnosti.",
        )

    decision_columns = list(zip(*df.apply(_rule_values, axis=1))) if len(df) else [[], [], [], [], []]
    df["recommended_action"] = pd.Series(decision_columns[0], index=df.index, dtype="string")
    df["suggested_parameter_to_check"] = pd.Series(decision_columns[1], index=df.index, dtype="string")
    df["decision_reason"] = pd.Series(decision_columns[2], index=df.index, dtype="string")
    df["confidence"] = pd.Series(decision_columns[3], index=df.index, dtype="string")
    df["validator_checklist"] = pd.Series(decision_columns[4], index=df.index, dtype="string")

    return df, warnings


def build_ai_cases_input(
    cases_enriched: pd.DataFrame,
    *,
    top_n: int = 100,
    exclude_stitok: bool = True,
    instruction_categories: list[str] | None = None,
    min_report_count: int = 0,
    product_match_status: list[str] | None = None,
    has_risk_flag: list[int] | None = None,
    trigger_detected: list[int] | None = None,
) -> pd.DataFrame:
    df = cases_enriched.copy()

    if product_match_status and "product_match_status" in df.columns:
        df = df[df["product_match_status"].isin(product_match_status)]
    if exclude_stitok and "instruction_category" in df.columns:
        df = df[df["instruction_category"] != "Štítok"]
    if instruction_categories and "instruction_category" in df.columns:
        df = df[df["instruction_category"].isin(instruction_categories)]
    if min_report_count and "report_count" in df.columns:
        df = df[df["report_count"] >= min_report_count]
    if has_risk_flag is not None and len(has_risk_flag) > 0 and "has_risk_flag" in df.columns:
        df = df[df["has_risk_flag"].isin(has_risk_flag)]
    if trigger_detected is not None and len(trigger_detected) > 0 and "trigger_detected" in df.columns:
        df = df[df["trigger_detected"].isin(trigger_detected)]

    if "report_count" in df.columns:
        df = df.sort_values(by=["report_count"], ascending=[False], kind="mergesort")
    if top_n is not None and top_n > 0:
        df = df.head(top_n)
    return df.reset_index(drop=True)


def build_summary(
    reports_clean: pd.DataFrame,
    cases_mvp: pd.DataFrame,
    cases_enriched: pd.DataFrame,
    ai_cases_input: pd.DataFrame,
) -> pd.DataFrame:
    matched = int((cases_enriched["product_match_status"] == "Produkt nájdený").sum())
    unmatched = int((cases_enriched["product_match_status"] != "Produkt nájdený").sum())

    summary_rows = [
        {"metric": "total_reports", "value": int(len(reports_clean))},
        {"metric": "total_cases", "value": int(len(cases_mvp))},
        {"metric": "matched_products_count", "value": matched},
        {"metric": "unmatched_products_count", "value": unmatched},
        {"metric": "total_ai_input_rows", "value": int(len(ai_cases_input))},
    ]

    category_pivot = (
        cases_enriched.groupby("instruction_category", dropna=False)
        .agg(
            report_count_sum=("report_count", "sum"),
            case_id_count=("case_id", "count"),
        )
        .reset_index()
    )

    summary_df = pd.DataFrame(summary_rows)
    summary_df["section"] = "overview"

    category_pivot["metric"] = "instruction_category"
    category_pivot["section"] = "pivot"

    return pd.concat([summary_df, category_pivot], ignore_index=True, sort=False)


def build_decision_summary(
    cases_enriched: pd.DataFrame,
    ai_cases_input: pd.DataFrame,
) -> pd.DataFrame:
    def _count_table(df: pd.DataFrame, group_cols: list[str], count_name: str) -> pd.DataFrame:
        available_cols = [column for column in group_cols if column in df.columns]
        if not available_cols:
            return pd.DataFrame(columns=["section", *group_cols, count_name])
        table = (
            df.groupby(available_cols, dropna=False)
            .size()
            .reset_index(name=count_name)
        )
        for column in group_cols:
            if column not in table.columns:
                table[column] = pd.NA
        table["section"] = "count"
        return table[["section", *group_cols, count_name]]

    action_counts = _count_table(cases_enriched, ["recommended_action"], "case_count")
    if not action_counts.empty:
        action_counts["metric"] = "recommended_action"

    confidence_counts = _count_table(cases_enriched, ["confidence"], "case_count")
    if not confidence_counts.empty:
        confidence_counts["metric"] = "confidence"

    category_action_counts = _count_table(
        cases_enriched,
        ["instruction_category", "recommended_action"],
        "case_count",
    )
    if not category_action_counts.empty:
        category_action_counts["metric"] = "instruction_category_recommended_action"

    top_20_columns = [
        column
        for column in [
            "case_id",
            "case_number",
            "KĂłd produktu",
            "PĹ™eloĹľenĂ© instrukce",
            "report_count",
            "instruction_category",
            "recommended_action",
            "suggested_parameter_to_check",
            "confidence",
            "product_match_status",
        ]
        if column in cases_enriched.columns
    ]
    top_20_cases = cases_enriched.sort_values(
        by=["report_count"],
        ascending=[False],
        kind="mergesort",
    ).head(20)
    if top_20_columns:
        top_20_cases = top_20_cases.loc[:, top_20_columns]
    top_20_cases = top_20_cases.copy()
    top_20_cases["section"] = "top_20_cases"
    top_20_cases["metric"] = "report_count"

    parts = [action_counts, confidence_counts, category_action_counts, top_20_cases]
    parts = [part for part in parts if not part.empty]
    if not parts:
        return pd.DataFrame(columns=["section", "metric"])

    return pd.concat(parts, ignore_index=True, sort=False)


def build_summary_for_demo(
    reports_clean: pd.DataFrame,
    cases_mvp: pd.DataFrame,
    cases_enriched: pd.DataFrame,
    ai_cases_input: pd.DataFrame,
    invalid_reports: pd.DataFrame,
) -> pd.DataFrame:
    matched = int((cases_enriched["product_match_status"] == "Produkt nĂˇjdenĂ˝").sum())
    unmatched = int((cases_enriched["product_match_status"] != "Produkt nĂˇjdenĂ˝").sum())

    summary_rows = [
        {"metric": "total_reports", "value": int(len(reports_clean)), "section": "overview"},
        {"metric": "total_cases", "value": int(len(cases_mvp)), "section": "overview"},
        {"metric": "matched_products_count", "value": matched, "section": "overview"},
        {"metric": "unmatched_products_count", "value": unmatched, "section": "overview"},
        {"metric": "invalid_reports_count", "value": int(len(invalid_reports)), "section": "overview"},
        {"metric": "total_ai_input_rows", "value": int(len(ai_cases_input)), "section": "overview"},
    ]

    category_pivot = (
        cases_enriched.groupby("instruction_category", dropna=False)
        .agg(
            report_count_sum=("report_count", "sum"),
            case_id_count=("case_id", "count"),
        )
        .reset_index()
    )
    category_pivot["metric"] = "instruction_category"
    category_pivot["section"] = "pivot"

    return pd.concat([pd.DataFrame(summary_rows), category_pivot], ignore_index=True, sort=False)


def build_decision_summary_for_demo(cases_enriched: pd.DataFrame) -> pd.DataFrame:
    total_cases = max(int(len(cases_enriched)), 1)

    def _count_table(group_cols: list[str], count_name: str) -> pd.DataFrame:
        available_cols = [column for column in group_cols if column in cases_enriched.columns]
        if not available_cols:
            return pd.DataFrame(columns=[*group_cols, count_name, "share_of_cases"])
        table = (
            cases_enriched.groupby(available_cols, dropna=False)
            .size()
            .reset_index(name=count_name)
        )
        for column in group_cols:
            if column not in table.columns:
                table[column] = pd.NA
        table["share_of_cases"] = table[count_name] / total_cases
        return table[[*group_cols, count_name, "share_of_cases"]]

    def _title_row(title: str) -> pd.DataFrame:
        return pd.DataFrame([{"section": title, "row_type": "section_title"}])

    def _data_block(title: str, frame: pd.DataFrame) -> pd.DataFrame:
        if frame.empty:
            return pd.DataFrame(columns=["section", "row_type"])
        block = frame.copy()
        block.insert(0, "section", title)
        block["row_type"] = "data"
        return block

    action_counts = _count_table(["recommended_action"], "case_count")
    if not action_counts.empty:
        action_counts = action_counts.sort_values(
            by=["case_count", "recommended_action"],
            ascending=[False, True],
            kind="mergesort",
        ).reset_index(drop=True)

    confidence_counts = _count_table(["confidence"], "case_count")
    if not confidence_counts.empty:
        confidence_counts = confidence_counts.sort_values(
            by=["case_count", "confidence"],
            ascending=[False, True],
            kind="mergesort",
        ).reset_index(drop=True)

    matrix_counts = _count_table(["instruction_category", "recommended_action"], "case_count")
    if not matrix_counts.empty:
        matrix_counts = matrix_counts.sort_values(
            by=["instruction_category", "case_count", "recommended_action"],
            ascending=[True, False, True],
            kind="mergesort",
        ).reset_index(drop=True)

    top_20_columns = [
        column
        for column in [
            "case_id",
            "case_number",
            "KĂłd produktu",
            "PĹ™eloĹľenĂ© instrukce",
            "report_count",
            "instruction_category",
            "recommended_action",
            "suggested_parameter_to_check",
            "confidence",
            "product_match_status",
        ]
        if column in cases_enriched.columns
    ]
    top_20_cases = cases_enriched.sort_values(
        by=["report_count", "last_reported_at"],
        ascending=[False, False],
        kind="mergesort",
    ).head(20)
    if top_20_columns:
        top_20_cases = top_20_cases.loc[:, top_20_columns]
    top_20_cases = top_20_cases.copy().reset_index(drop=True)
    top_20_cases.insert(0, "rank", range(1, len(top_20_cases) + 1))
    top_20_cases.insert(0, "section", "Top 20 cases by report_count")
    top_20_cases["row_type"] = "data"
    top_20_cases = top_20_cases.reindex(
        columns=[
            "section",
            "row_type",
            "rank",
            "case_id",
            "case_number",
            "KĂłd produktu",
            "PĹ™eloĹľenĂ© instrukce",
            "report_count",
            "instruction_category",
            "recommended_action",
            "suggested_parameter_to_check",
            "confidence",
            "product_match_status",
        ]
    )

    parts = [
        _title_row("Overview by recommended_action"),
        _data_block("Overview by recommended_action", action_counts),
        _title_row("Overview by confidence"),
        _data_block("Overview by confidence", confidence_counts),
        _title_row("Matrix by instruction_category and recommended_action"),
        _data_block("Matrix by instruction_category and recommended_action", matrix_counts),
        _title_row("Top 20 cases by report_count"),
        top_20_cases,
    ]
    parts = [part for part in parts if not part.empty]
    return pd.concat(parts, ignore_index=True, sort=False) if parts else pd.DataFrame(columns=["section", "row_type"])


def build_invalid_reports_for_demo(reports_clean: pd.DataFrame) -> pd.DataFrame:
    df = reports_clean.copy()
    required_fields = ["report_id", "KĂłd produktu", "PĹ™eloĹľenĂ© instrukce"]
    available_fields = [column for column in required_fields if column in df.columns]
    if not available_fields:
        invalid = df.head(0).copy()
        invalid["invalid_reason"] = pd.Series(dtype="string")
        return invalid

    invalid_mask = df[available_fields].isna().any(axis=1)
    invalid = df.loc[invalid_mask].copy()

    def _reason(row: pd.Series) -> str:
        missing = [column for column in required_fields if column in row.index and pd.isna(row[column])]
        return "; ".join(missing) if missing else "unknown"

    invalid["invalid_reason"] = invalid.apply(_reason, axis=1)
    return invalid.reset_index(drop=True)


def _format_workbook_for_demo(workbook) -> None:
    header_font = Font(bold=True)
    section_font = Font(bold=True)
    section_fill = SECTION_FILL
    header_fill = HEADER_FILL
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    header_alignment = Alignment(wrap_text=True, vertical="center")

    for worksheet in workbook.worksheets:
        if worksheet.max_row < 1 or worksheet.max_column < 1:
            continue

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        header_row = worksheet[1]
        for cell in header_row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        section_row_indices: set[int] = set()
        header_names = [worksheet.cell(row=1, column=col).value for col in range(1, worksheet.max_column + 1)]
        row_type_col_index = None
        if "row_type" in header_names:
            row_type_col_index = header_names.index("row_type") + 1

        if row_type_col_index is not None:
            for row_idx in range(2, worksheet.max_row + 1):
                row_type = worksheet.cell(row=row_idx, column=row_type_col_index).value
                if row_type == "section_title":
                    section_row_indices.add(row_idx)
                    for col_idx in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.font = section_font
                        cell.fill = section_fill
                        cell.alignment = header_alignment

        for col_idx, header in enumerate(header_names, start=1):
            if header is None:
                continue
            if header == "row_type":
                width = 12
            elif header in LONG_TEXT_COLUMN_WIDTHS:
                width = LONG_TEXT_COLUMN_WIDTHS[header]
            else:
                values = []
                for row_idx in range(2, worksheet.max_row + 1):
                    if row_idx in section_row_indices:
                        continue
                    value = worksheet.cell(row=row_idx, column=col_idx).value
                    if value is not None:
                        values.append(len(str(value)))
                width = max([len(str(header))] + values + [10]) + 2
                width = min(width, 45)
            worksheet.column_dimensions[get_column_letter(col_idx)].width = width

        for row_idx in range(2, worksheet.max_row + 1):
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.value is None:
                    continue
                if isinstance(cell.value, str) and len(cell.value) > 40:
                    cell.alignment = wrap_alignment

        if "row_type" in header_names:
            row_type_index = header_names.index("row_type") + 1
            for row_idx in range(2, worksheet.max_row + 1):
                row_type = worksheet.cell(row=row_idx, column=row_type_index).value
                if row_type == "section_title":
                    worksheet.row_dimensions[row_idx].height = 20


@dataclass
class PipelineResult:
    reports_clean: pd.DataFrame
    products_clean: pd.DataFrame
    cases_mvp: pd.DataFrame
    cases_enriched: pd.DataFrame
    ai_cases_input: pd.DataFrame
    invalid_reports: pd.DataFrame
    summary: pd.DataFrame
    decision_summary: pd.DataFrame
    warnings: list[str]


def run_pipeline(
    reports_raw: pd.DataFrame,
    products_raw: pd.DataFrame,
    *,
    top_n: int = 100,
    exclude_stitok: bool = True,
    instruction_categories: list[str] | None = None,
    min_report_count: int = 0,
    product_match_status: list[str] | None = None,
    has_risk_flag: list[int] | None = None,
    trigger_detected: list[int] | None = None,
) -> PipelineResult:
    warnings: list[str] = []

    reports_clean, report_warnings = prepare_reports_clean(reports_raw)
    warnings.extend(report_warnings)

    products_clean, product_warnings = prepare_products_clean(products_raw)
    warnings.extend(product_warnings)

    cases_mvp = group_cases_mvp(reports_clean)
    cases_enriched = build_cases_enriched(cases_mvp, products_clean)
    cases_enriched, decision_warnings = build_decision_maker(cases_enriched)
    warnings.extend(decision_warnings)
    invalid_reports = build_invalid_reports_for_demo(reports_clean)
    ai_cases_input = build_ai_cases_input(
        cases_enriched,
        top_n=top_n,
        exclude_stitok=exclude_stitok,
        instruction_categories=instruction_categories,
        min_report_count=min_report_count,
        product_match_status=product_match_status,
        has_risk_flag=has_risk_flag,
        trigger_detected=trigger_detected,
    )
    summary = build_summary_for_demo(reports_clean, cases_mvp, cases_enriched, ai_cases_input, invalid_reports)
    decision_summary = build_decision_summary_for_demo(cases_enriched)

    return PipelineResult(
        reports_clean=reports_clean,
        products_clean=products_clean,
        cases_mvp=cases_mvp,
        cases_enriched=cases_enriched,
        ai_cases_input=ai_cases_input,
        invalid_reports=invalid_reports,
        summary=summary,
        decision_summary=decision_summary,
        warnings=warnings,
    )


def export_workbook(result: PipelineResult) -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        sheets_in_order = [
            ("summary", result.summary),
            ("decision_summary", result.decision_summary),
            ("ai_cases_input", result.ai_cases_input),
            ("cases_enriched", result.cases_enriched),
            ("cases_mvp", result.cases_mvp),
            ("invalid_reports", result.invalid_reports),
            ("reports_clean", result.reports_clean),
            ("products_clean", result.products_clean),
        ]
        for sheet_name, df in sheets_in_order:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        _format_workbook_for_demo(writer.book)
    return buffer.getvalue()

